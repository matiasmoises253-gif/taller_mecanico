import os
import sys, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

datos = json.loads(open(os.path.join(os.path.dirname(__file__), 'reporte_datos.json')).read())
resumen = datos['resumen']
recientes = datos['recientes']
por_semana = datos['porSemana']
por_estado = datos.get('porEstado', {})
desglose = datos.get('desglose')

wb = Workbook()

AZUL = '003366'
AZUL_FILL = PatternFill('solid', fgColor=AZUL)
GRIS_FILL = PatternFill('solid', fgColor='F2F2F2')
VERDE = '008000'
# El texto "S/." va entre comillas para que Excel no confunda la "S" con el
# código de "segundos" y termine mostrando los montos como si fueran horas/fechas.
FORMATO_MONEDA = '"S/. "#,##0.00'
BORDE = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

def header_cell(ws, row, col, texto, width=None):
    c = ws.cell(row=row, column=col, value=texto)
    c.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    c.fill = AZUL_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = BORDE
    if width: ws.column_dimensions[get_column_letter(col)].width = width
    return c

def data_cell(ws, row, col, valor, bold=False, color='000000', fmt=None, fill=None):
    c = ws.cell(row=row, column=col, value=valor)
    c.font = Font(bold=bold, color=color, name='Arial', size=10)
    c.alignment = Alignment(vertical='center', wrap_text=True)
    c.border = BORDE
    if fmt: c.number_format = fmt
    if fill: c.fill = fill
    return c

# ── HOJA 1: Dashboard ──────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Dashboard'
ws1.row_dimensions[1].height = 35
ws1.merge_cells('A1:B1')
t = ws1.cell(1,1, 'REPORTE DE RENDIMIENTO - MULTISERVICIOS CÁRDENAS')
t.font = Font(bold=True, color='FFFFFF', name='Arial', size=14)
t.fill = AZUL_FILL
t.alignment = Alignment(horizontal='center', vertical='center')

header_cell(ws1, 2, 1, 'Indicador', 25)
header_cell(ws1, 2, 2, 'Valor', 20)

kpis = [
    ('Ingresos Totales', f"S/. {float(resumen['total']):,.2f}"),
    ('Órdenes de Servicio', resumen['total_ordenes']),
    ('Órdenes Finalizadas', resumen['completadas']),
    ('Ticket Promedio', f"S/. {float(resumen['ticket_promedio']):,.2f}"),
]
for i, (k, v) in enumerate(kpis):
    fill = GRIS_FILL if i % 2 == 0 else None
    data_cell(ws1, i+3, 1, k, bold=True, fill=fill)
    data_cell(ws1, i+3, 2, v, bold=True, color=VERDE, fill=fill)

# ── HOJA 2: Órdenes ───────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Órdenes de Servicio')
headers = ['Fecha','Cliente','Vehículo','Descripción','Estado','Monto (S/.)']
widths =  [14,     18,       20,        35,            14,      14]
for c, (h, w) in enumerate(zip(headers, widths), 1):
    header_cell(ws2, 1, c, h, w)

for i, r in enumerate(recientes):
    fill = GRIS_FILL if i % 2 == 0 else None
    data_cell(ws2, i+2, 1, r['fecha'], fill=fill)
    data_cell(ws2, i+2, 2, r['cliente'], fill=fill)
    data_cell(ws2, i+2, 3, r['vehiculo'], fill=fill)
    data_cell(ws2, i+2, 4, r['descripcion'], fill=fill)
    data_cell(ws2, i+2, 5, r['estado'], fill=fill)
    data_cell(ws2, i+2, 6, float(r['monto']), bold=True, color=VERDE, fmt=FORMATO_MONEDA, fill=fill)

# ── HOJA 3: Ingresos por Semana ───────────────────────────────────────────────
ws3 = wb.create_sheet('Ingresos por Semana')
header_cell(ws3, 1, 1, 'Periodo', 18)
header_cell(ws3, 1, 2, 'Ingresos (S/.)', 20)
total = 0
for i, (sem, val) in enumerate(por_semana.items()):
    fill = GRIS_FILL if i % 2 == 0 else None
    data_cell(ws3, i+2, 1, sem, fill=fill)
    data_cell(ws3, i+2, 2, float(val), fmt=FORMATO_MONEDA, fill=fill)
    total += float(val)
row = len(por_semana) + 2
data_cell(ws3, row, 1, 'TOTAL', bold=True)
data_cell(ws3, row, 2, total, bold=True, color=VERDE, fmt=FORMATO_MONEDA)

# ── HOJA 4: Distribución por Estado ───────────────────────────────────────────
ws4 = wb.create_sheet('Distribución por Estado')
header_cell(ws4, 1, 1, 'Estado', 18)
header_cell(ws4, 1, 2, 'Órdenes', 14)
header_cell(ws4, 1, 3, 'Porcentaje', 14)
total_estados = sum(por_estado.values()) or 1
for i, (estado, cnt) in enumerate(por_estado.items()):
    fill = GRIS_FILL if i % 2 == 0 else None
    data_cell(ws4, i+2, 1, estado, fill=fill)
    data_cell(ws4, i+2, 2, cnt, fill=fill)
    data_cell(ws4, i+2, 3, cnt/total_estados, fmt='0%', fill=fill)

# ── HOJA 5: Resumen Económico (Ingresos, Gastos y Ganancia Neta) ─────────────
ws5 = wb.create_sheet('Resumen Económico')
header_cell(ws5, 1, 1, 'Concepto', 34)
header_cell(ws5, 1, 2, 'Monto (S/.)', 20)

fila_actual = 2
if desglose:
    data_cell(ws5, fila_actual, 1, 'Ingresos por Mano de Obra', bold=True, color=VERDE); data_cell(ws5, fila_actual, 2, float(desglose['mano_obra']), color=VERDE, fmt=FORMATO_MONEDA); fila_actual += 1
    data_cell(ws5, fila_actual, 1, 'Ingresos por Venta de Repuestos', bold=True, color=VERDE); data_cell(ws5, fila_actual, 2, float(desglose['repuestos_ingreso']), color=VERDE, fmt=FORMATO_MONEDA); fila_actual += 1
    data_cell(ws5, fila_actual, 1, 'Total de Ingresos', bold=True, fill=GRIS_FILL); data_cell(ws5, fila_actual, 2, float(desglose['total_ingresos']), bold=True, color=VERDE, fmt=FORMATO_MONEDA, fill=GRIS_FILL); fila_actual += 1
    data_cell(ws5, fila_actual, 1, 'Gastos por Compra de Repuestos (Inventario)', bold=True, color='dc2626'); data_cell(ws5, fila_actual, 2, float(desglose['gastos_inventario']), color='dc2626', fmt=FORMATO_MONEDA); fila_actual += 1
    ganancia = float(desglose['ganancia_neta'])
    color_ganancia = VERDE if ganancia >= 0 else 'dc2626'
    fill_ganancia = PatternFill('solid', fgColor='D9F2D9' if ganancia >= 0 else 'FADBD8')
    data_cell(ws5, fila_actual, 1, 'GANANCIA NETA', bold=True, fill=fill_ganancia); data_cell(ws5, fila_actual, 2, ganancia, bold=True, color=color_ganancia, fmt=FORMATO_MONEDA, fill=fill_ganancia); fila_actual += 1
    fila_actual += 1

data_cell(ws5, fila_actual, 1, 'Ticket Promedio', bold=True); data_cell(ws5, fila_actual, 2, float(resumen['ticket_promedio']), fmt=FORMATO_MONEDA); fila_actual += 1
data_cell(ws5, fila_actual, 1, 'Órdenes Totales', bold=True); data_cell(ws5, fila_actual, 2, resumen['total_ordenes']); fila_actual += 1
data_cell(ws5, fila_actual, 1, 'Órdenes Finalizadas/Completadas', bold=True); data_cell(ws5, fila_actual, 2, resumen['completadas'])

wb.save(os.path.join(os.path.dirname(__file__), 'reporte_output.xlsx'))
print('OK')
